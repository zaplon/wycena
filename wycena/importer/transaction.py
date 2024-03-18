import csv
import datetime
import logging
import uuid

import googlemaps
import openpyxl
from pydantic import ValidationError
from pydantic.v1 import DateError
from pydantic.v1.datetime_parse import parse_date
from sqlmodel import Session

from wycena import abstract_models
from wycena.db import models as db_models
from wycena.db.models import get_engine
from wycena.settings import settings


def perform_import(file_path: str, **defaults):
    rows = get_rows(file_path)
    with Session(get_engine()) as session:
        for row in rows:
            upsert_transaction(session, row, **defaults)
        session.commit()


def get_rows(file_path) -> [dict]:
    extension = file_path.split('.')[-1]
    if extension in ['xlsx', 'xls']:
        wb = openpyxl.load_workbook(file_path)
        rows = []
        for sheet in wb.worksheets:
            first_row = next(sheet.iter_rows())
            headers = [c.value for c in first_row]
            for r in sheet.iter_rows():
                rows.append(dict(zip(headers, r.values())))
        return rows
    elif extension == 'csv':
        with open(file_path, "r") as f:
            dialect = csv.Sniffer().sniff(f.read(), delimiters=';,| ')
            f.seek(0)
            rows = list(csv.DictReader(f, dialect=dialect))
        return rows
    else:
        raise ValueError(f"Not supported extension: {extension}")


def upsert_transaction(session, row: dict, **defaults):
    row = {k.lower(): v for k, v in row.items()}
    data = {}
    for field, col_handler in FIELD_HANDLERS.items():
        for col_name in col_handler[0]:
            if col_name in row:
                try:
                    val = col_handler[1](row[col_name]) if col_handler[1] is not None else row[col_name]
                except ValueError as e:
                    logging.error(e)
                    return
                data[field] = val
                break
    data['long'], data['lat'] = get_coords(data)
    for k, v in defaults.items():
        if not data.get(k):
            data[k] = v
    data['id'] = uuid.uuid4()
    try:
        t = abstract_models.Transaction(**data)
    except ValidationError as e:
        logging.warning(e)
        return
    session.add(db_models.Transaction(**t.model_dump()))


def handle_float(val):
    return float(
        val.replace(",", ".").replace(" ", "").replace(u'\xa0', "")
    ) if val else None


def handle_int(val):
    return int(val) if val else None


def handle_date(val):
    try:
        date = parse_date(val) if val else None
    except DateError:
        fmts = ["%d.%m.%Y"]
        for fmt in fmts:
            try:
                date = datetime.datetime.strptime(val, fmt).date()
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Could not parse date: {val}")
    return datetime.datetime.combine(date, datetime.time.min)


def get_coords(data) -> (float, float):
    address = f"{data['street']} {data['building_nr']} {data['city']}"
    geocode_result = get_maps_client().geocode(address)
    return geocode_result


def get_maps_client() -> googlemaps.Client:
    return googlemaps.Client(key=settings.MAPS_API_KEY)


FIELD_HANDLERS = {
    "street": (["ulica"], None),
    "building_nr": (["numer", "nr bud.", "numer budynku"], None),
    "city": (["miejscowość", "miasto"], None),
    "district": (["dzielnica"], None),
    "floor": (["piętro", "pietro", "Kond."], handle_int),
    "price": (["cena", "cena trans. [zł]"], handle_float),
    "area": (["powierzchnia", "p.u. [m2]"], handle_float),
    "transaction_date": (["data transakcji"], handle_date),
    "primary_market": (["rynek", "rodzaj rynku"], lambda x: True if x in ["1", "pierwotny"] else False),
    "number_of_rooms": (["izby", "Liczba pokoi"], handle_int),
    "number_of_floors": (["liczba pięter"], handle_int)
}
